"""Export service — generate MIS Excel report with 5 sheets."""

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import select, func, case
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.ndc_record import NdcRecord
from app.models.ndc_approval import NdcApproval


HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header_row(ws, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


async def generate_mis_excel(db: AsyncSession) -> io.BytesIO:
    """Generate MIS report with 5 sheets:
    1. All Records
    2. Pending Summary
    3. Stage-wise Bottlenecks
    4. Department Clearance
    5. Approver Load
    """
    wb = Workbook()

    # --- Sheet 1: All Records ---
    ws1 = wb.active
    ws1.title = "All Records"
    headers = [
        "Person Number", "Employee Name", "Business Unit", "Legal Employer",
        "Location", "Location City", "Department", "NDC Stage",
        "Resignation Date", "Last Working Date", "NDC Assigned Date",
        "NDC Initiated Date", "NDC Completed Date", "Pending Days",
    ]
    ws1.append(headers)
    _style_header_row(ws1, len(headers))

    records_q = await db.execute(
        select(NdcRecord).order_by(NdcRecord.ndc_initiated_date.desc())
    )
    records = records_q.scalars().all()
    today = date.today()
    for r in records:
        pending = (today - r.ndc_initiated_date).days if r.ndc_initiated_date and r.ndc_stage != "NDC Completed" else None
        ws1.append([
            r.person_number, r.employee_name, r.business_unit, r.legal_employer,
            r.location, r.location_city, r.department_reporting_name, r.ndc_stage,
            r.resignation_date, r.last_working_date, r.ndc_assigned_date,
            r.ndc_initiated_date, r.ndc_completed_date, pending,
        ])

    # Auto-width
    for col in ws1.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 3, 35)

    # --- Sheet 2: Pending Summary ---
    ws2 = wb.create_sheet("Pending Summary")
    ws2.append(["NDC Stage", "Count"])
    _style_header_row(ws2, 2)

    stage_q = await db.execute(
        select(NdcRecord.ndc_stage, func.count().label("cnt"))
        .group_by(NdcRecord.ndc_stage)
        .order_by(func.count().desc())
    )
    for r in stage_q.all():
        ws2.append([r.ndc_stage, r.cnt])

    # --- Sheet 3: Stage-wise Bottlenecks ---
    ws3 = wb.create_sheet("Stage Bottlenecks")
    ws3.append(["Stage", "Pending Count"])
    _style_header_row(ws3, 2)

    bottleneck_q = await db.execute(
        select(NdcApproval.stage_name, func.count().label("cnt"))
        .where(NdcApproval.status == "PENDING")
        .group_by(NdcApproval.stage_name)
        .order_by(func.count().desc())
    )
    for r in bottleneck_q.all():
        ws3.append([r.stage_name, r.cnt])

    # --- Sheet 4: Department Clearance ---
    ws4 = wb.create_sheet("Dept Clearance")
    ws4.append(["Department", "Total", "Completed", "Clearance %"])
    _style_header_row(ws4, 4)

    dept_q = await db.execute(
        select(
            NdcRecord.department_reporting_name,
            func.count().label("total"),
            func.sum(case((NdcRecord.ndc_stage == "NDC Completed", 1), else_=0)).label("completed"),
        )
        .group_by(NdcRecord.department_reporting_name)
        .order_by(func.count().desc())
    )
    for r in dept_q.all():
        total = r.total or 0
        completed = r.completed or 0
        pct = round(100.0 * completed / total, 1) if total > 0 else 0
        ws4.append([r.department_reporting_name, total, completed, pct])

    # --- Sheet 5: Approver Load ---
    ws5 = wb.create_sheet("Approver Load")
    ws5.append(["Approver", "Stage", "Pending Count"])
    _style_header_row(ws5, 3)

    approver_q = await db.execute(
        select(
            NdcApproval.approver_name,
            NdcApproval.stage_name,
            func.count().label("cnt"),
        )
        .where(NdcApproval.status == "PENDING")
        .group_by(NdcApproval.approver_name, NdcApproval.stage_name)
        .order_by(func.count().desc())
    )
    for r in approver_q.all():
        ws5.append([r.approver_name, r.stage_name, r.cnt])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
