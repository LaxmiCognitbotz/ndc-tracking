import datetime
import io
import logging
import os
from pathlib import Path

import httpx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from datetime import date
from app.models.ndc_record import NdcRecord
from app.models.upload_batch import UploadBatch
from app.services.common_service import _propagate_department_dates
from app.services.ingest_service import ingest_excel_file
from app.services.sharepoint_service import SharePointService, get_httpx_client

logger = logging.getLogger(__name__)

class SharePointSyncService:
    def __init__(self):
        self.sharepoint = SharePointService()
        self.upload_dir = Path(os.getenv("UPLOAD_DIR", "uploads"))
        self.reports_subfolder = os.getenv("SHAREPOINT_REPORTS_SUBFOLDER", "NDC_Reports")

    async def check_and_ingest_new_files(self, db: AsyncSession) -> dict:
        """
        Check the SharePoint reports folder for new Excel files,
        download them, and ingest them into the database.
        Returns a summary of the sync operation.
        """
        sync_results = {
            "status": "success",
            "files_found": 0,
            "files_downloaded": 0,
            "files_ingested": 0,
            "details": [],
            "errors": []
        }

        async with get_httpx_client(timeout=60.0) as client:
            try:
                # 1. Resolve site ID
                site_id = await self.sharepoint.get_site_id(client)
                
                # 2. Get base drive and target folder details
                drive_id, base_folder_path = await self.sharepoint.get_drive_details(client, site_id)
                
                # 3. Construct the reports subfolder path
                if base_folder_path:
                    reports_path = f"{base_folder_path}/{self.reports_subfolder}"
                else:
                    reports_path = self.reports_subfolder
                
                clean_path = "/".join([p for p in reports_path.split("/") if p])
                encoded_path = self.sharepoint._encode_path_segments(clean_path)
                
                # 4. List files inside the reports folder
                url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives/{drive_id}/root:/{encoded_path}:/children"
                logger.info(f"Checking SharePoint directory: {clean_path}")
                
                token = await self.sharepoint.get_access_token()
                headers = {"Authorization": f"Bearer {token}"}
                
                response = await client.get(url, headers=headers)
                if response.status_code != 200:
                    err_msg = f"Failed to list SharePoint folder '{clean_path}': {response.status_code} - {response.text}"
                    logger.error(err_msg)
                    sync_results["status"] = "failed"
                    sync_results["errors"].append(err_msg)
                    return sync_results

                children = response.json().get("value", [])
                # Filter for Excel files only
                excel_files = []
                for item in children:
                    if "file" in item:
                        name = item.get("name", "")
                        ext = Path(name).suffix.lower()
                        if ext in (".xlsx", ".xlsb", ".xls"):
                            excel_files.append(item)
                
                sync_results["files_found"] = len(excel_files)
                logger.info(f"Found {len(excel_files)} Excel file(s) in SharePoint folder '{clean_path}'.")

                for file_item in excel_files:
                    file_name = file_item.get("name")
                    logger.info(f"Processing SharePoint file: '{file_name}'")

                    # Check if this file has already been successfully/partially ingested in the database
                    query = select(UploadBatch).where(
                        UploadBatch.file_name == file_name,
                        UploadBatch.status.in_(["success", "partial"])
                    )
                    db_result = await db.execute(query)
                    existing_batch = db_result.scalar_one_or_none()

                    if existing_batch:
                        msg = f"File '{file_name}' already ingested (Batch ID: {existing_batch.id}, Status: {existing_batch.status}). Skipping."
                        logger.info(msg)
                        sync_results["details"].append({"file_name": file_name, "status": "skipped", "message": msg})
                        continue

                    # If not ingested yet, download the file
                    download_url = file_item.get("@microsoft.graph.downloadUrl")
                    if not download_url:
                        err_msg = f"File '{file_name}' does not have a download URL from SharePoint."
                        logger.error(err_msg)
                        sync_results["errors"].append(err_msg)
                        sync_results["details"].append({"file_name": file_name, "status": "failed", "message": err_msg})
                        continue

                    # Ensure the local upload directory exists
                    self.upload_dir.mkdir(parents=True, exist_ok=True)
                    local_file_path = self.upload_dir / file_name

                    logger.info(f"Downloading '{file_name}' to local path '{local_file_path}'...")
                    dl_response = await client.get(download_url)
                    if dl_response.status_code != 200:
                        err_msg = f"Failed to download '{file_name}' from SharePoint (status {dl_response.status_code})."
                        logger.error(err_msg)
                        sync_results["errors"].append(err_msg)
                        sync_results["details"].append({"file_name": file_name, "status": "failed", "message": err_msg})
                        continue

                    # Save the file content locally
                    with open(local_file_path, "wb") as f:
                        f.write(dl_response.content)
                    
                    sync_results["files_downloaded"] += 1
                    logger.info(f"Downloaded '{file_name}' successfully. Initiating ingestion...")

                    # Run database ingestion
                    try:
                        ingest_result = await ingest_excel_file(
                            file_path=local_file_path,
                            file_name=file_name,
                            uploaded_by="sharepoint_sync",
                            db=db,
                            source_type="sharepoint"
                        )
                        
                        sync_results["files_ingested"] += 1
                        msg = f"Ingested '{file_name}' successfully (Batch ID: {ingest_result['batch_id']}, Status: {ingest_result['status']})."
                        logger.info(msg)
                        sync_results["details"].append({
                            "file_name": file_name,
                            "status": "ingested",
                            "batch_id": ingest_result["batch_id"],
                            "records_processed": ingest_result["records_processed"],
                            "records_failed": ingest_result["records_failed"],
                            "message": msg
                        })
                    except Exception as ingest_err:
                        err_msg = f"Error ingesting file '{file_name}': {str(ingest_err)}"
                        logger.exception(err_msg)
                        sync_results["errors"].append(err_msg)
                        sync_results["details"].append({"file_name": file_name, "status": "failed", "message": err_msg})
                    finally:
                        # Clean up local file after ingestion (success or fail)
                        if local_file_path.exists():
                            try:
                                os.remove(local_file_path)
                                logger.info(f"Cleaned up local file: '{local_file_path}'")
                            except Exception as cleanup_err:
                                logger.error(f"Failed to delete local file '{local_file_path}': {cleanup_err}")

            except Exception as e:
                err_msg = f"Failed to execute SharePoint sync: {str(e)}"
                logger.exception(err_msg)
                sync_results["status"] = "failed"
                sync_results["errors"].append(err_msg)

        return sync_results

    async def sync_fnf_completed_records(self, db: AsyncSession) -> dict:
        """
        Scan SharePoint F&F Documents folder for employee directories (named with person numbers),
        and update their is_fnf_completed status in the database to True.
        """
        results = {
            "status": "success",
            "folders_found": 0,
            "records_updated": 0,
            "errors": []
        }
        
        async with get_httpx_client(timeout=60.0) as client:
            try:
                # 1. Resolve site ID
                site_id = await self.sharepoint.get_site_id(client)
                
                # 2. Get base drive and target folder details
                drive_id, base_folder_path = await self.sharepoint.get_drive_details(client, site_id)
                
                # 3. Construct the F&F documents path
                fnf_folder = "F&F_Documents"
                if base_folder_path:
                    fnf_path = f"{base_folder_path}/{fnf_folder}"
                else:
                    fnf_path = fnf_folder
                
                clean_path = "/".join([p for p in fnf_path.split("/") if p])
                encoded_path = self.sharepoint._encode_path_segments(clean_path)
                
                # 4. List children in the F&F folder
                url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives/{drive_id}/root:/{encoded_path}:/children"
                logger.info(f"Checking SharePoint F&F directory: {clean_path}")
                
                token = await self.sharepoint.get_access_token()
                headers = {"Authorization": f"Bearer {token}"}
                
                response = await client.get(url, headers=headers)
                if response.status_code != 200:
                    err_msg = f"Failed to list SharePoint F&F folder '{clean_path}': {response.status_code} - {response.text}"
                    logger.error(err_msg)
                    results["status"] = "failed"
                    results["errors"].append(err_msg)
                    return results

                children = response.json().get("value", [])
                
                # 5. Extract employee person numbers (numeric folder names or files)
                person_numbers = []
                for item in children:
                    name = item.get("name", "")
                    base_name = name.split(".")[0]
                    if base_name.isdigit():
                        person_numbers.append(int(base_name))
                
                results["folders_found"] = len(person_numbers)
                logger.info(f"Found {len(person_numbers)} F&F folder(s)/file(s) in SharePoint.")
                
                # 6. Query and update records
                # Update matching records to True
                completed_count = 0
                if person_numbers:
                    stmt = select(NdcRecord).where(
                        NdcRecord.person_number.in_(person_numbers),
                        NdcRecord.is_fnf_completed == False
                    )
                    db_result = await db.execute(stmt)
                    records_to_update = db_result.scalars().all()
                    
                    today = date.today()
                    for record in records_to_update:
                        record.is_fnf_completed = True
                        if not record.fnf_completed_date:
                            record.fnf_completed_date = today
                        record.is_fnf_revision = False
                        
                        # Also propagate dates
                        await _propagate_department_dates(record.id, today, db)
                        completed_count += 1
                
                # Update non-matching records to False
                if person_numbers:
                    stmt_revert = select(NdcRecord).where(
                        NdcRecord.person_number.notin_(person_numbers),
                        NdcRecord.is_fnf_completed == True
                    )
                else:
                    stmt_revert = select(NdcRecord).where(
                        NdcRecord.is_fnf_completed == True
                    )
                db_result_revert = await db.execute(stmt_revert)
                records_to_revert = db_result_revert.scalars().all()
                
                reverted_count = 0
                for record in records_to_revert:
                    # Never revert a record that has been manually closed — is_fnf_closed is permanent
                    if record.is_fnf_closed:
                        continue
                    record.is_fnf_completed = False
                    record.fnf_completed_date = None
                    reverted_count += 1
                    
                if completed_count > 0 or reverted_count > 0:
                    await db.commit()
                    logger.info(f"F&F Completed Sync: Completed {completed_count}, Reverted {reverted_count} records.")
                
                results["records_updated"] = completed_count
                results["records_reverted"] = reverted_count
                        
            except Exception as e:
                err_msg = f"Failed to execute SharePoint F&F Completed Sync: {str(e)}"
                logger.exception(err_msg)
                results["status"] = "failed"
                results["errors"].append(err_msg)
                
        return results

    async def generate_and_upload_fnf_closed_report(self, db: AsyncSession) -> dict:
        """
        READ-ONLY: Queries person numbers where ndc_stage='NDC Completed' and is_fnf_closed=False,
        generates an Excel with just those person numbers,
        and uploads it to SharePoint folder 'fnf_closed_report'.
        Does NOT write, update, or delete anything in the database.
        """
        results = {
            "status": "success",
            "records_exported": 0,
            "uploaded_file_name": None,
            "errors": [],
        }

        try:
            # ── 1. Read-only SELECT — only person numbers where ndc_stage='NDC Completed' and is_fnf_closed=False ─

            stmt = select(NdcRecord.person_number).where(
                NdcRecord.ndc_stage == "NDC Completed",
                NdcRecord.is_fnf_closed == False
            ).order_by(NdcRecord.person_number)
            db_result = await db.execute(stmt)
            person_numbers = [row[0] for row in db_result.fetchall()]
            results["records_exported"] = len(person_numbers)
            logger.info(f"FNF Report: Found {len(person_numbers)} record(s) with ndc_stage='NDC Completed' and is_fnf_closed=False.")

            # ── 2. Build Excel workbook (Person Number only) ─────────────────────
            HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
            HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
            HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
            THIN_BORDER = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            wb = Workbook()
            ws = wb.active
            ws.title = "FNF Active Records"

            # Single column — Person Number only
            ws.append(["Person Number"])
            header_cell = ws.cell(row=1, column=1)
            header_cell.font = HEADER_FONT
            header_cell.fill = HEADER_FILL
            header_cell.alignment = HEADER_ALIGNMENT
            header_cell.border = THIN_BORDER

            for pn in person_numbers:
                ws.append([pn])

            ws.column_dimensions["A"].width = 20

            # Save to in-memory buffer
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            file_bytes = buf.read()

            # ── 3. Upload to SharePoint (timestamped, replaces previous) ─────────
            now_str = datetime.datetime.now().strftime("%d_%B_%Y_%I.%M%p")
            file_name = f"FNF_Active_Report_{now_str}.xlsx"
            results["uploaded_file_name"] = file_name

            upload_subfolder = "F&F_Active_Report"

            async with get_httpx_client(timeout=120.0) as client:
                # Resolve site & drive
                site_id = await self.sharepoint.get_site_id(client)
                drive_id, base_folder_path = await self.sharepoint.get_drive_details(client, site_id)

                token = await self.sharepoint.get_access_token()
                headers = {
                    "Authorization": f"Bearer {token}",
                    "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                }

                # Build the folder path
                if base_folder_path:
                    folder_path = f"{base_folder_path}/{upload_subfolder}"
                else:
                    folder_path = upload_subfolder
                clean_folder = "/".join([p for p in folder_path.split("/") if p])
                encoded_folder = self.sharepoint._encode_path_segments(clean_folder)

                # Delete all existing files in the folder before uploading
                list_url = (
                    f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives/{drive_id}"
                    f"/root:/{encoded_folder}:/children"
                )
                auth_header = {"Authorization": f"Bearer {token}"}
                list_resp = await client.get(list_url, headers=auth_header)
                logger.info(f"FNF Active Report: List folder status={list_resp.status_code} for '{clean_folder}'")
                if list_resp.status_code == 200:
                    existing_items = list_resp.json().get("value", [])
                    logger.info(f"FNF Active Report: Found {len(existing_items)} item(s) in folder, cleaning up old .xlsx files...")
                    for item in existing_items:
                        if "file" in item and item.get("name", "").endswith(".xlsx"):
                            item_id = item["id"]
                            item_name = item.get("name", "")
                            del_url = (
                                f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives/{drive_id}"
                                f"/items/{item_id}"
                            )
                            del_resp = await client.delete(del_url, headers=auth_header)
                            if del_resp.status_code in (200, 204):
                                logger.info(f"FNF Active Report: Deleted old file '{item_name}' from SharePoint.")
                            else:
                                logger.warning(
                                    f"FNF Active Report: Failed to delete '{item_name}' "
                                    f"(status {del_resp.status_code}): {del_resp.text}"
                                )
                else:
                    logger.warning(
                        f"FNF Active Report: Could not list folder '{clean_folder}' "
                        f"(status {list_resp.status_code}): {list_resp.text} — skipping cleanup."
                    )

                # Upload the new file
                dest_path = f"{clean_folder}/{file_name}"
                encoded_path = self.sharepoint._encode_path_segments(dest_path)
                upload_url = (
                    f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives/{drive_id}"
                    f"/root:/{encoded_path}:/content"
                )

                logger.info(f"FNF Active Report: Uploading '{file_name}' to '{dest_path}'...")
                response = await client.put(upload_url, headers=headers, content=file_bytes)

                if response.status_code in (200, 201):
                    logger.info(
                        f"FNF Active Report: Successfully uploaded '{file_name}' "
                        f"({results['records_exported']} records)."
                    )
                else:
                    err_msg = (
                        f"FNF Active Report: Upload failed for '{file_name}'. "
                        f"Status {response.status_code}: {response.text}"
                    )
                    logger.error(err_msg)
                    results["status"] = "failed"
                    results["errors"].append(err_msg)

        except Exception as e:
            err_msg = f"FNF Active Report: Unexpected error — {str(e)}"
            logger.exception(err_msg)
            results["status"] = "failed"
            results["errors"].append(err_msg)

        return results
